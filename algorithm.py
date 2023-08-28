from openpyxl.writer.excel import save_workbook
import openpyxl
import csv

filepathxl = ''


# def set_name(filepath, newName):
#     indexFileName = filepath.rfind('/') + 1
#     if indexFileName != 0:
#         fileName = filepath[:indexFileName:]
#         filepath = fileName + newName + '.xlsx'
#     return filepath


def txt_to_exel(filepath):
    global filepathxl
    with open(filepath, "rt", newline='', encoding='UTF-8') as file_obj:
        list1 = [i for i in file_obj]
        for row in file_obj.read():
            list1.append(row)
        list1 = [list(','.join(row.split('\t')) for row in list1)]
        wb = openpyxl.Workbook()
        ws = wb.active
        for row in list1:
            for cell in row:
                cell = list((cell.replace('"', '').split(';')))
                ws.append(cell)
        wb.save(filepathxl)
    return filepathxl


def csv_to_excel(filepath):
    global filepathxl
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(filepath, 'r', newline='', encoding='utf-8') as file_obj:
        reader = csv.reader(file_obj, delimiter=';')
        for row in reader:
            ws.append(row)

    wb.save(filepathxl)
    return (filepathxl)


def open_filexl(filepath, entry):
    global filepathxl
    data = entry

    try:
        wb = openpyxl.load_workbook(filepath)
    except:
        wb = openpyxl.Workbook()

    sheet = wb.active  # Получение активного листа

    listquality = []
    for cell in sheet['D']:
        if cell.value not in listquality:
            listquality.append(cell.value)
    listquality.pop(0)
    if None in listquality:
        listquality.remove(None)
    for quality in listquality:
        if quality in wb.sheetnames:
            wb.remove(wb[quality])
        ws = wb.create_sheet(str(quality))
        flagname = True
        for row in wb.active.iter_rows(values_only=True):
            if row[1] is None or row[3] is None:
                continue
            if data in row[1] and quality in row[3] or flagname:
                ws.append(row)
                flagname = False
    if 'Other' in wb.sheetnames:
        wb.remove(wb['Other'])
    wsother = wb.create_sheet('Other')
    if 'Empty' in wb.sheetnames:
        wb.remove(wb['Empty'])
    wsempty = wb.create_sheet('Empty')
    flagname = True
    for row in wb.active.iter_rows(values_only=True):
        if flagname:
            wsempty.append(row)
            wsother.append(row)
            flagname = False
        elif row[1] is None and row[3] is None:
            wsother.append(row)
        elif data in row[1] and row[3] is None:
            wsempty.append(row)
        elif data not in row[1]:
            wsother.append(row)
    for sheep in wb.sheetnames:
        if wb[sheep].max_row <= 1:
            wb.remove(wb[sheep])
    save_workbook(wb, filepathxl)
